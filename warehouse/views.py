from django.shortcuts import render
from .models import Inventory
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Product, Warehouse, Inventory, Transaction, UserProfile
from django.http import HttpResponse, JsonResponse  # 添加JsonResponse导入
import csv
from django.db.models import Q, F  # 添加F表达式导入
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage  # 补充导入
from django.db import transaction  # 新增导入
from .forms import ProductImportForm
from django.contrib import messages  # 新增导入
from django.utils import timezone
# 添加Decimal导入
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl import load_workbook  # 新增这行导入

def check_purchase_permission(user):
    return user.userprofile.permissions in ['admin', 'purchase']

def check_sales_permission(user):
    return user.userprofile.permissions in ['admin', 'sales']

@login_required
@user_passes_test(check_purchase_permission)
def stock_in(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 获取所有产品数据
                materials = request.POST.getlist('material_number')
                products = request.POST.getlist('product_name')
                specs = request.POST.getlist('spec')
                suppliers = request.POST.getlist('supplier')
                prices = request.POST.getlist('purchase_price')
                warehouses = request.POST.getlist('warehouse_name')
                quantities = request.POST.getlist('quantity')

                for i in range(len(materials)):
                    # 数据校验
                    if not all([materials[i], products[i], specs[i], suppliers[i], prices[i], warehouses[i], quantities[i]]):
                        raise ValueError("第{}行存在空字段".format(i+1))

                    # 创建/获取产品
                    product, created = Product.objects.get_or_create(
                        material_number=materials[i],
                        defaults={
                            'name': products[i],
                            'spec': specs[i],
                            'supplier': suppliers[i],
                            'purchase_price': prices[i]
                        }
                    )

                    # 创建仓库
                    warehouse, _ = Warehouse.objects.get_or_create(
                        name=warehouses[i],
                        defaults={'location': '默认位置', 'capacity': 1000}
                    )

                    # 更新库存
                    inventory, _ = Inventory.objects.update_or_create(
                        product=product,
                        warehouse=warehouse,
                        defaults={'quantity': F('quantity') + int(quantities[i])}
                    )

                    # 创建交易记录
                    Transaction.objects.create(
                        transaction_type='IN',
                        product=product,
                        warehouse=warehouse,
                        quantity=quantities[i],
                        operator=request.user
                    )

                return redirect('inventory_query')

        except Exception as e:
            error = str(e)
            return render(request, 'warehouse/stock_in.html', {
                'error': error,
                # 新增保留已填数据
                'post_data': request.POST 
            })
    # 新增数据回填处理
    post_data = request.session.pop('post_data', None)
    return render(request, 'warehouse/stock_in.html', {'post_data': post_data})

@login_required
@user_passes_test(check_sales_permission)
def stock_out(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                materials = request.POST.getlist('material_number')
                specs = request.POST.getlist('spec')
                suppliers = request.POST.getlist('supplier')
                warehouse_ids = request.POST.getlist('warehouse')
                quantities = request.POST.getlist('quantity')
                sale_prices = request.POST.getlist('sale_price')

                for i in range(len(materials)):
                    product = get_object_or_404(
                        Product,
                        material_number=materials[i],
                        spec=specs[i],
                        supplier=suppliers[i]
                    )
                    
                    warehouse = get_object_or_404(Warehouse, id=warehouse_ids[i])
                    quantity = int(quantities[i])
                    sale_price = Decimal(sale_prices[i])

                    inventory = Inventory.objects.get(
                        product=product,
                        warehouse=warehouse
                    )
                    
                    if inventory.quantity < quantity:
                        raise ValueError(f"{product.name} 出库数量超过库存数量")
                    
                    inventory.quantity = F('quantity') - quantity
                    inventory.save()

                    Transaction.objects.create(
                        transaction_type='OUT',
                        product=product,
                        warehouse=warehouse,
                        quantity=quantity,
                        operator=request.user,
                        sale_price=sale_price
                    )
                
                return redirect('inventory_query')
                
        except Inventory.DoesNotExist:  # 新增库存不存在异常处理
            error = "指定仓库中不存在该产品库存"
        except Exception as e:
            error = str(e)
            return render(request, 'warehouse/stock_out.html', {
                'error': error,
                'products': Product.objects.all(),
                'warehouses': Warehouse.objects.all(),
                'specs': Product.objects.values_list('spec', flat=True).distinct(),
                'suppliers': Product.objects.values_list('supplier', flat=True).distinct()
            })
    
    return render(request, 'warehouse/stock_out.html', {
        'products': Product.objects.all(),
        'warehouses': Warehouse.objects.all(),
        'specs': Product.objects.values_list('spec', flat=True).distinct(),
        'suppliers': Product.objects.values_list('supplier', flat=True).distinct()
    })

@login_required(login_url='/login/')  # 添加登录跳转配置
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='管理员').exists() or u.userprofile.permissions == 'admin', login_url='/login/')  # 增加login_url参数
def user_list(request):
    search_query = request.GET.get('search', '')  # 新增这行获取搜索参数
    users = User.objects.select_related('userprofile').all().order_by('-date_joined')
    
    if search_query:
        users = users.filter(
            Q(id__icontains=search_query) |
            Q(username__icontains=search_query)
        )
    
    return render(request, 'warehouse/user_list.html', {
        'users': users,
        'search_query': search_query
    })

@user_passes_test(lambda u: u.is_superuser)
def user_create(request):
    if request.method == 'POST':
        permissions = request.POST.get('permissions')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not all([username, password]):
            return render(request, 'warehouse/user_form.html', {
                'error': '用户名和密码必须填写',
                'user': None
            })
            
        try:
            # 创建用户时不再设置is_superuser
            user = User.objects.create_user(
                username=username,
                password=password,
                is_superuser=False  # 强制关闭超级用户权限
            )
            UserProfile.objects.create(
                user=user,
                permissions=permissions
            )
            return redirect('user_list')
        except Exception as e:
            return render(request, 'warehouse/user_form.html', {
                'error': str(e),
                'user': None
            })
    else:
        return render(request, 'warehouse/user_form.html', {'user': None})

@user_passes_test(lambda u: u.is_superuser)
def user_edit(request, user_id):
    user = get_object_or_404(User.objects.select_related('userprofile'), id=user_id)
    if request.method == 'POST':
        permissions = request.POST.get('permissions')
        username = request.POST.get('username')
        password = request.POST.get('password')
        permissions = request.POST.get('permissions')  # 新增权限字段
        
        try:
            user.username = username
            if password:
                user.set_password(password)
                
            # 更新用户权限配置
            user.userprofile.permissions = permissions  # 新增这行
            user.userprofile.save()  # 新增这行
            
            user.save()
            return redirect('user_list')
        except Exception as e:
            return render(request, 'warehouse/user_form.html', {
                'error': str(e),
                'user': user
            })
    else:
        # 添加UserProfile存在性检查
        if not hasattr(user, 'userprofile'):
            UserProfile.objects.create(user=user, permissions='sales')  # 默认值
        return render(request, 'warehouse/user_form.html', {
            'user': user,
            'current_permissions': user.userprofile.permissions
        })

@user_passes_test(lambda u: u.is_superuser)
def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')
    return render(request, 'warehouse/user_confirm_delete.html', {'user': user})


@user_passes_test(lambda u: u.is_authenticated and (u.is_superuser or u.groups.filter(name='管理员').exists() or u.userprofile.permissions == 'admin'), login_url='/login/')
def product_import_template(request):
    # 生成Excel模板
    wb = Workbook()
    ws = wb.active
    ws.title = "产品模板"
    # 移除仓库名称和初始数量列
    ws.append(['物料号', '产品名称', '产品型号', '供应商', '采购单价'])
    
    # 使用BytesIO代替save_virtual_workbook
    io = BytesIO()
    wb.save(io)
    io.seek(0)
    
    response = HttpResponse(
        io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_template.xlsx"'
    return response

@user_passes_test(lambda u: u.is_authenticated and (u.is_superuser or u.groups.filter(name='管理员').exists() or u.userprofile.permissions == 'admin'), login_url='/login/')
def product_import(request):
    if request.method == 'POST' and 'excel_submit' in request.POST:
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                errors = []
                success_count = 0
                
                with transaction.atomic():
                    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        sid = transaction.savepoint()
                        try:
                            # 新增列数验证（只需要前5列）
                            if len(row) < 5:
                                raise ValueError("文件列数不足，请使用正确的产品导入模板")
                                
                            # 只取前五列数据
                            material_number, name, spec, supplier, purchase_price, *_ = row
                            # 新增产品存在性验证
                            if Product.objects.filter(material_number=material_number).exists():
                                existing_product = Product.objects.get(material_number=material_number)
                                # 拦截新增已存在产品
                                raise ValueError(f"物料号 {material_number} 已存在（产品名称：{existing_product.name}）")
                                
                            # 新增必填字段验证
                            if not all(row[0:5]):
                                raise ValueError("前五列（物料号到采购单价）不能为空")
                                
                            warehouse, _ = Warehouse.objects.get_or_create(
                                name='默认仓库',  # 固定使用默认仓库
                                defaults={'location': '默认仓库', 'capacity': 1000}
                            )
                            
                            product = Product.objects.create(
                                material_number=material_number,
                                name=row[1],
                                spec=row[2],
                                supplier=row[3],
                                purchase_price=row[4],
                                warehouse=warehouse,
                                sale_price=0,
                                quantity=0  # 初始数量设为0
                            )
                            
                            Inventory.objects.update_or_create(
                                product=product,
                                warehouse=warehouse,
                                defaults={'quantity': 0}  # 固定为0
                            )
                            success_count += 1
                        except Exception as e:
                            error_parts = str(e).split(":")
                            error_type = error_parts[0] if len(error_parts) > 1 else "数据错误"
                            error_msg = error_parts[-1].strip()
                            errors.append({
                                "row": row_index,
                                "type": error_type,
                                "message": error_msg
                            })
                            transaction.savepoint_rollback(sid)
                            
                # 处理导入结果
                if errors:
                    messages.error(request, f"成功导入{success_count}条，失败{len(errors)}条")
                else:
                    messages.success(request, f"成功导入{success_count}条产品信息记录")
                
                request.session['import_errors'] = errors
                request.session['success_count'] = success_count  # 新增这行
                return redirect('import_result')

            except Exception as e:
                messages.error(request, f"文件解析失败: {str(e)}")
                return redirect('import_result')
    
    return render(request, 'warehouse/import.html')  # 确保模板文件路径正确

# 将import_result函数移到product_import函数外
@login_required(login_url='/login/')
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='管理员').exists() or u.userprofile.permissions == 'admin', login_url='/login/')
def import_result(request):
    errors = request.session.get('import_errors', [])
    success_count = request.session.get('success_count', 0)  # 添加默认值
    return render(request, 'warehouse/import_result.html', {
        'import_errors': errors,
        'success_count': success_count  # 传递变量到模板
    })
@login_required
def inventory_query(request):
    search_field = request.GET.get('field', 'all')
    search_query = request.GET.get('search', '')
    
    # 构建查询条件
    filter_condition = Q()
    if search_query:
        if search_field == 'material_number':
            filter_condition |= Q(product__material_number__icontains=search_query)
        elif search_field == 'name':
            filter_condition |= Q(product__name__icontains=search_query)
        elif search_field == 'spec':
            filter_condition |= Q(product__spec__icontains=search_query)
        elif search_field == 'supplier':
            filter_condition |= Q(product__supplier__icontains=search_query)
        else:  # 处理全字段搜索
            filter_condition |= (
                Q(product__material_number__icontains=search_query) |
                Q(product__name__icontains=search_query) |
                Q(product__spec__icontains=search_query) |
                Q(product__supplier__icontains=search_query)
            )
    
    inventories = Inventory.objects.select_related('product', 'warehouse').filter(filter_condition)
    
    # 分页处理
    paginator = Paginator(inventories, 25)  # 每页25条
    page_number = request.GET.get('page')
    
    try:
        inventories = paginator.page(page_number)
    except PageNotAnInteger:
        inventories = paginator.page(1)
    except EmptyPage:
        inventories = paginator.page(paginator.num_pages)

    return render(request, 'warehouse/inventory_query.html', {
        'inventories': inventories,
        'search_field': search_field,
        'search_query': search_query,
        'paginator': paginator
    })

@login_required
def get_products_by_field(request):
    field = request.GET.get('field')
    value = request.GET.get('value')
    
    try:
        product = Product.objects.get(**{field: value})
        return JsonResponse({
            'exists': True,
            'product': {
                'material_number': product.material_number,
                'name': product.name,
                'spec': product.spec,
                'supplier': product.supplier,
                'purchase_price': str(product.purchase_price),
                'warehouse': product.warehouse.name  # 确保返回仓库名称
            }
        })
    except Product.DoesNotExist:
        # 模糊搜索时限制字段范围
        allowed_fields = ['material_number', 'name', 'spec', 'supplier']
        if field not in allowed_fields:
            return JsonResponse({'error': 'Invalid search field'}, status=400)
            
        results = Product.objects.filter(**{f'{field}__icontains': value})[:10]
        return JsonResponse({
            'exists': False,
            'results': [
                {
                    'material_number': p.material_number,
                    'spec': p.spec,
                    'name': p.name,
                    'supplier': p.supplier
                } for p in results
            ]
        }, safe=False)
    except Product.MultipleObjectsReturned:
        return JsonResponse({'error': '找到多个匹配产品'}, status=400)

# 新增导出功能
@login_required
@user_passes_test(lambda u: u.is_superuser or u.userprofile.permissions == 'admin')
def export_inventory(request):
    selected_ids = request.GET.get('ids', '')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['物料号', '产品名称', '型号', '供应商', '采购单价', '销售单价', '仓库', '数量'])
    
    if selected_ids:
        inventories = Inventory.objects.filter(id__in=selected_ids.split(','))
    else:
        inventories = Inventory.objects.all()
    
    for inv in inventories.select_related('product', 'warehouse'):
        writer.writerow([
            inv.product.material_number,
            inv.product.name,
            inv.product.spec,
            inv.product.supplier,
            inv.product.purchase_price,
            inv.product.sale_price,
            inv.warehouse.name,
            inv.quantity
        ])
    
    return response

# 在Product模型中新增查询方法
@staticmethod
def get_by_field(field, value):
    try:
        return Product.objects.get(**{field: value})
    except Product.DoesNotExist:
        return None
    except Product.MultipleObjectsReturned:
        raise
    
    @staticmethod
    def search_by_field(field, value):
        return Product.objects.filter(**{f'{field}__icontains': value})
def is_admin(user):
    return user.groups.filter(name='管理员').exists() or user.userprofile.permissions == 'admin'


@login_required
def export_print_queue(request):
    # 获取打印队列数据
    selected_ids = request.GET.get('ids', '')
    search_query = request.GET.get('search', '')
    search_field = request.GET.get('field', 'all')
    
    # 创建Excel响应
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="inventory_print.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "库存清单"
    
    # 添加表头（包含搜索条件）
    ws.append(['搜索条件：', f"{search_query} ({search_field})"])
    headers = ['物料号', '产品名称', '型号', '数量', '所在仓库']
    ws.append(headers)
    
    # 获取数据
    inventories = Inventory.objects.filter(id__in=selected_ids.split(',')) if selected_ids else Inventory.objects.all()
    
    for inv in inventories.select_related('product', 'warehouse'):
        ws.append([
            inv.product.material_number,
            inv.product.name,
            inv.product.spec,
            inv.quantity,
            inv.warehouse.name
        ])
    
    wb.save(response)
    return response

@user_passes_test(check_purchase_permission)
def stock_import_template(request):
    # 生成库存导入专用模板
    wb = Workbook()
    ws = wb.active
    ws.title = "库存入库模板"
    ws.append(['物料号', '产品名称', '产品型号', '供应商', '采购单价', '仓库名称', '入库数量'])
    
    io = BytesIO()
    wb.save(io)
    io.seek(0)
    
    response = HttpResponse(
        io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_import_template.xlsx"'
    return response

@login_required
@user_passes_test(check_purchase_permission)
def stock_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                errors = []
                success_count = 0
                
                with transaction.atomic():
                    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        sid = transaction.savepoint()
                        try:
                            # 新增字段校验
                            if not all(row[0:5]):
                                raise ValueError("前五列（物料号到采购单价）不能为空")
                                
                            # 检查产品匹配
                            product = Product.objects.filter(
                                material_number=row[0],
                                name=row[1],
                                spec=row[2],
                                supplier=row[3],
                                purchase_price=row[4]
                            ).first()
                            
                            if not product:
                                raise ValueError("未找到匹配的产品信息")
                                
                            # 处理仓库逻辑
                            warehouse_name = row[5]
                            if product.warehouse.name == '默认仓库':
                                warehouse, _ = Warehouse.objects.update_or_create(
                                    name=warehouse_name,
                                    defaults={'location': warehouse_name, 'capacity': 1000}
                                )
                            else:
                                warehouse, _ = Warehouse.objects.get_or_create(
                                    name=warehouse_name,
                                    defaults={'location': warehouse_name, 'capacity': 1000}
                                )
                            
                            # 更新库存逻辑修改
                            quantity = int(row[6])
                            inventory, created = Inventory.objects.get_or_create(
                                product=product,
                                warehouse=warehouse,
                                defaults={'quantity': quantity}  # 新增时直接赋值
                            )
                            
                            if not created:
                                # 已有记录时才使用F表达式更新
                                inventory.quantity = F('quantity') + quantity
                                inventory.save()
                            # 创建交易记录
                            Transaction.objects.create(
                                transaction_type='IN',
                                product=product,
                                warehouse=warehouse,
                                quantity=quantity,
                                operator=request.user
                            )
                            
                            success_count += 1
                            
                        except Exception as e:
                            error_parts = str(e).split(":")
                            error_type = error_parts[0] if len(error_parts) > 1 else "数据错误"
                            error_msg = error_parts[-1].strip()
                            errors.append({
                                "row": row_index,
                                "type": error_type,
                                "message": error_msg
                            })
                            transaction.savepoint_rollback(sid)
                            
                # 处理导入结果
                # 处理结果提示
                messages.success(request, f"成功导入{success_count}条记录")
                if errors:
                    messages.warning(request, f"失败{len(errors)}条记录，请查看错误详情")
                
                request.session['import_errors'] = errors
                request.session['success_count'] = success_count
                return redirect('stock_import_result')

            except Exception as e:
                messages.error(request, f"文件解析失败: {str(e)}")
                return redirect('stock_import_result')
    
    return render(request, 'warehouse/stock_import.html')

@login_required
@user_passes_test(check_purchase_permission)
def stock_import_result(request):
    errors = request.session.get('import_errors', [])
    success_count = request.session.get('success_count', 0)
    
    # 添加消息处理
    if errors:
        messages.error(request, f"成功导入{success_count}条，失败{len(errors)}条")
    else:
        messages.success(request, f"成功导入{success_count}条库存记录")
        
    return render(request, 'warehouse/stock_import_result.html', {
        'import_errors': errors,
        'success_count': success_count
    })

@login_required
@user_passes_test(lambda u: u.is_superuser or u.userprofile.permissions == 'admin')
def get_print_data(request):
    selected_ids = request.GET.get('ids', '').split(',')
    inventories = Inventory.objects.filter(id__in=selected_ids).select_related('product', 'warehouse')
    
    data = [{
        'material_number': inv.product.material_number,
        'name': inv.product.name,
        'spec': inv.product.spec,
        'warehouse': inv.warehouse.name,
        'quantity': inv.quantity,
        'purchase_price': str(inv.product.purchase_price),
        'sale_price': str(inv.product.sale_price)
    } for inv in inventories]
    
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(check_sales_permission)
def stock_out_bulk(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        bulk_errors = []
        success_count = 0
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            with transaction.atomic():
                for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    sid = transaction.savepoint()
                    try:
                        # 字段校验
                        if len(row) < 7:
                            raise ValueError("文件列数不足，请使用正确的批量出库模板")
                        
                        material_number = row[0] or ''
                        name = row[1] or ''
                        spec = row[2] or ''
                        supplier = row[3] or ''
                        sale_price = Decimal(row[4] or 0)
                        warehouse_name = row[5] or ''
                        quantity = int(row[6] or 0)
                        
                        # 获取产品和库存
                        try:
                            product = Product.objects.get(
                                material_number=material_number,
                                name=name,
                                spec=spec,
                                supplier=supplier
                            )
                            warehouse = Warehouse.objects.get(name=warehouse_name)
                            inventory = Inventory.objects.get(product=product, warehouse=warehouse)
                        except Product.DoesNotExist:
                            raise ValueError("库存中不存在该产品")
                        except Warehouse.DoesNotExist:
                            raise ValueError("仓库不存在")
                        except Inventory.DoesNotExist:
                            raise ValueError("该仓库中无此产品库存")
                        
                        # 数量校验
                        if quantity <= 0:
                            raise ValueError("出库数量必须大于0")
                        if inventory.quantity < quantity:
                            raise ValueError(f"库存不足（当前库存：{inventory.quantity}）")
                        
                        # 价格校验
                        if sale_price <= product.purchase_price:
                            raise ValueError(f"销售单价必须大于采购单价（当前采购价：{product.purchase_price}）")
                        
                        # 更新销售价格
                        product.sale_price = sale_price
                        product.save()
                        
                        # 库存检查
                        inventory = Inventory.objects.get(product=product, warehouse=warehouse)
                        if inventory.quantity < quantity:
                            raise ValueError("库存不足")
                            
                        # 执行出库
                        inventory.quantity = F('quantity') - quantity
                        inventory.save()
                        
                        # 创建交易记录
                        Transaction.objects.create(
                            transaction_type='OUT',
                            product=product,
                            warehouse=warehouse,
                            quantity=quantity,
                            operator=request.user,
                            sale_price=sale_price
                        )
                        success_count += 1
                    except Exception as e:
                        bulk_errors.append({
                            'row': row_index,
                            'message': f"第{row_index}行错误：{str(e)}"
                        })
                        transaction.savepoint_rollback(sid)
            
            messages.success(request, f'成功处理{success_count}条记录，失败{len(bulk_errors)}条')
            request.session['bulk_errors'] = bulk_errors
            return redirect('stock_out_result')

        except Exception as e:
            messages.error(request, f'文件处理失败: {str(e)}')
            return redirect('stock_out_result')
    
    return render(request, 'warehouse/stock_out_bulk.html')

def stock_out_result(request):
    bulk_errors = request.session.get('bulk_errors', [])
    return render(request, 'warehouse/stock_out_result.html', {
        'bulk_errors': bulk_errors
    })
def stock_out_template(request):
    # 生成Excel模板
    wb = Workbook()
    ws = wb.active
    ws.append(['物料号', '产品名称', '产品型号', '供应商', '销售单价', '仓库', '数量'])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bulk_stock_out_template.xlsx"'
    return response

@user_passes_test(lambda u: u.is_superuser or u.userprofile.permissions == 'admin')
def get_print_data(request):
    selected_ids = request.GET.get('ids', '').split(',')
    inventories = Inventory.objects.filter(id__in=selected_ids).select_related('product', 'warehouse')
    
    data = [{
        'material_number': inv.product.material_number,
        'name': inv.product.name,
        'spec': inv.product.spec,
        'warehouse': inv.warehouse.name,
        'quantity': inv.quantity,
        'purchase_price': str(inv.product.purchase_price),
        'sale_price': str(inv.product.sale_price)
    } for inv in inventories]
    
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(check_sales_permission)
def stock_out_result(request):
    errors = request.session.get('bulk_errors', [])
    success_count = request.session.get('success_count', 0)
    return render(request, 'warehouse/stock_out_result.html', {
        'bulk_errors': errors,
        'success_count': success_count
    })
